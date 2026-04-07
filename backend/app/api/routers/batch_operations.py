"""API router for batch PEC operations (Journees entreprise)."""

import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.tenant_context import TenantContext, get_tenant_context
from app.db.session import get_db
from app.domain.schemas.batch_operation import (
    BatchCreateRequest,
    BatchOperationResponse,
    BatchPecResult,
    BatchSummaryResponse,
    MarketingCodeResponse,
)
from app.services import batch_operation_service

router = APIRouter(prefix="/api/v1/batch", tags=["batch"])


@router.get(
    "/marketing-codes",
    response_model=list[MarketingCodeResponse],
    summary="Lister les codes marketing disponibles",
    description=(
        "Retourne tous les tags Cosium avec le nombre de clients associes. "
        "Filtrage optionnel par date (derniere facture ou creation client)."
    ),
)
def list_marketing_codes(
    date_from: date | None = Query(
        None, description="Date debut (filtre derniere facture ou creation client)"
    ),
    date_to: date | None = Query(
        None, description="Date fin (filtre derniere facture ou creation client)"
    ),
    db: Session = Depends(get_db),
    tenant_ctx: TenantContext = Depends(get_tenant_context),
) -> list[MarketingCodeResponse]:
    return batch_operation_service.get_available_marketing_codes(
        db,
        tenant_id=tenant_ctx.tenant_id,
        date_from=date_from,
        date_to=date_to,
    )


@router.post(
    "/create",
    response_model=BatchOperationResponse,
    status_code=201,
    summary="Creer une operation batch",
    description=(
        "Cree une operation batch pour tous les clients lies a un code marketing. "
        "Filtrage optionnel par date pour restreindre aux clients actifs sur une periode."
    ),
)
def create_batch(
    payload: BatchCreateRequest,
    db: Session = Depends(get_db),
    tenant_ctx: TenantContext = Depends(get_tenant_context),
) -> BatchOperationResponse:
    return batch_operation_service.create_batch(
        db,
        tenant_id=tenant_ctx.tenant_id,
        marketing_code=payload.marketing_code,
        label=payload.label,
        user_id=tenant_ctx.user_id,
        date_from=payload.date_from,
        date_to=payload.date_to,
    )


@router.post(
    "/{batch_id}/process",
    response_model=BatchOperationResponse,
    summary="Traiter un batch (consolidation + pre-controle)",
    description=(
        "Lance la consolidation et le pre-controle pour chaque client du batch. "
        "Si async=true et que le batch depasse 50 clients, le traitement est lance "
        "en arriere-plan via Celery et le batch est retourne avec status='en_cours'."
    ),
)
def process_batch(
    batch_id: int,
    use_async: bool = Query(
        False, alias="async", description="Lancer en arriere-plan pour les gros batches"
    ),
    db: Session = Depends(get_db),
    tenant_ctx: TenantContext = Depends(get_tenant_context),
) -> BatchOperationResponse:
    batch = batch_operation_service.get_batch_by_id(
        db, tenant_ctx.tenant_id, batch_id
    )

    if use_async and batch.total_clients > 50:
        from app.tasks.batch_tasks import process_batch_async

        process_batch_async.delay(
            tenant_ctx.tenant_id, batch_id, tenant_ctx.user_id
        )
        return batch

    return batch_operation_service.process_batch(
        db,
        tenant_id=tenant_ctx.tenant_id,
        batch_id=batch_id,
        user_id=tenant_ctx.user_id,
    )


@router.post(
    "/{batch_id}/prepare-pec",
    response_model=BatchPecResult,
    summary="Preparer les PEC pour les clients prets",
    description="Cree une preparation PEC pour chaque client avec le statut 'pret'.",
)
def prepare_batch_pec(
    batch_id: int,
    db: Session = Depends(get_db),
    tenant_ctx: TenantContext = Depends(get_tenant_context),
) -> BatchPecResult:
    return batch_operation_service.prepare_batch_pec(
        db,
        tenant_id=tenant_ctx.tenant_id,
        batch_id=batch_id,
        user_id=tenant_ctx.user_id,
    )


@router.get(
    "/{batch_id}",
    response_model=BatchSummaryResponse,
    summary="Detail d'un batch avec tous les items",
    description="Retourne le detail complet d'une operation batch avec le statut de chaque client.",
)
def get_batch_detail(
    batch_id: int,
    db: Session = Depends(get_db),
    tenant_ctx: TenantContext = Depends(get_tenant_context),
) -> BatchSummaryResponse:
    return batch_operation_service.get_batch_summary(
        db, tenant_id=tenant_ctx.tenant_id, batch_id=batch_id
    )


@router.get(
    "",
    summary="Lister les operations batch",
    description="Liste paginee de toutes les operations batch du tenant.",
)
def list_batches(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db),
    tenant_ctx: TenantContext = Depends(get_tenant_context),
) -> dict:
    return batch_operation_service.list_batches(
        db,
        tenant_id=tenant_ctx.tenant_id,
        page=page,
        page_size=page_size,
    )


@router.get(
    "/{batch_id}/export",
    summary="Exporter le resume batch en Excel",
    description="Telecharge un fichier Excel avec le detail de l'operation batch.",
)
def export_batch(
    batch_id: int,
    db: Session = Depends(get_db),
    tenant_ctx: TenantContext = Depends(get_tenant_context),
) -> StreamingResponse:
    enriched = batch_operation_service.get_batch_summary_enriched(
        db, tenant_id=tenant_ctx.tenant_id, batch_id=batch_id
    )

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return _export_csv(enriched)

    return _export_excel(enriched)


def _export_excel(enriched: dict) -> StreamingResponse:
    """Generate a professional 2-sheet Excel export of batch summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    bold_font = Font(bold=True)
    green_fill = PatternFill(
        start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
    )
    red_fill = PatternFill(
        start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
    )
    amber_fill = PatternFill(
        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
    )

    batch = enriched["batch"]
    items = enriched["items"]

    # ── Sheet 1: Resume ──
    ws1 = wb.active
    ws1.title = "Resume"

    info_rows = [
        ("Operation batch", f"#{batch.id}"),
        ("Code marketing", batch.marketing_code),
        ("Label", batch.label or "-"),
        ("Statut", batch.status),
        ("Date debut", batch.started_at.strftime("%d/%m/%Y %H:%M") if batch.started_at else "-"),
        ("Date fin", batch.completed_at.strftime("%d/%m/%Y %H:%M") if batch.completed_at else "-"),
        ("", ""),
        ("Total clients", batch.total_clients),
        ("Clients prets", batch.clients_prets),
        ("Clients incomplets", batch.clients_incomplets),
        ("Clients en conflit", batch.clients_en_conflit),
        ("Clients en erreur", batch.clients_erreur),
    ]
    for row_idx, (label, value) in enumerate(info_rows, 1):
        c_label = ws1.cell(row=row_idx, column=1, value=label)
        c_label.font = bold_font
        c_label.border = thin_border
        c_val = ws1.cell(row=row_idx, column=2, value=value)
        c_val.border = thin_border
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40

    # ── Sheet 2: Detail clients ──
    ws2 = wb.create_sheet("Detail clients")
    headers = [
        "Client",
        "Telephone",
        "Email",
        "N. Secu",
        "Mutuelle",
        "Score (%)",
        "Statut",
        "Erreurs",
        "Alertes",
        "PEC ID",
        "Message erreur",
        "Traite le",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_fills = {
        "pret": green_fill,
        "incomplet": amber_fill,
        "conflit": red_fill,
        "erreur": red_fill,
    }

    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.get("customer_name", ""),
            item.get("phone", ""),
            item.get("email", ""),
            item.get("social_security_number", ""),
            item.get("mutuelle_name", ""),
            round(item.get("completude_score", 0), 1),
            item.get("status", ""),
            item.get("errors_count", 0),
            item.get("warnings_count", 0),
            item.get("pec_preparation_id") or "",
            item.get("error_message") or "",
            item.get("processed_at", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
        # Color the status cell
        status_cell = ws2.cell(row=row_idx, column=7)
        fill = status_fills.get(item.get("status"))
        if fill:
            status_cell.fill = fill

    # Column widths
    widths = [30, 18, 30, 18, 25, 10, 14, 10, 10, 10, 40, 20]
    for i, w in enumerate(widths):
        ws2.column_dimensions[chr(65 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"batch_{batch.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_csv(enriched: dict) -> StreamingResponse:
    """Fallback CSV export."""
    import csv

    batch = enriched["batch"]
    items = enriched["items"]

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([
        "Client", "Telephone", "Email", "N. Secu", "Mutuelle",
        "Score", "Statut", "Erreurs", "Alertes", "PEC ID", "Erreur",
    ])
    for item in items:
        writer.writerow([
            item.get("customer_name", ""),
            item.get("phone", ""),
            item.get("email", ""),
            item.get("social_security_number", ""),
            item.get("mutuelle_name", ""),
            round(item.get("completude_score", 0), 1),
            item.get("status", ""),
            item.get("errors_count", 0),
            item.get("warnings_count", 0),
            item.get("pec_preparation_id") or "",
            item.get("error_message") or "",
        ])

    content = buf.getvalue().encode("utf-8-sig")
    filename = f"batch_{batch.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
