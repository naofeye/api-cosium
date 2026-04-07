"""Client import service -- CSV/Excel file import and template generation."""
import csv
import io
import re

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.domain.schemas.clients import ClientImportError, ClientImportResult
from app.models.client import Customer
from app.services import audit_service

logger = get_logger("client_import_service")

# Column name mapping: French variants -> internal field names
_COLUMN_MAP: dict[str, str] = {
    # last_name
    "nom": "last_name",
    "nom de famille": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    # first_name
    "prenom": "first_name",
    "prénom": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    # email
    "email": "email",
    "e-mail": "email",
    "mail": "email",
    "adresse email": "email",
    "adresse e-mail": "email",
    # phone
    "telephone": "phone",
    "téléphone": "phone",
    "tel": "phone",
    "tél": "phone",
    "phone": "phone",
    "portable": "phone",
    "mobile": "phone",
    # birth_date
    "date de naissance": "birth_date",
    "date_naissance": "birth_date",
    "naissance": "birth_date",
    "birth_date": "birth_date",
    # address
    "adresse": "address",
    "address": "address",
    # city
    "ville": "city",
    "city": "city",
    # postal_code
    "code postal": "postal_code",
    "code_postal": "postal_code",
    "cp": "postal_code",
    "postal_code": "postal_code",
    # social_security_number
    "n° sécu": "social_security_number",
    "n° secu": "social_security_number",
    "numero secu": "social_security_number",
    "numéro sécu": "social_security_number",
    "nir": "social_security_number",
    "securite sociale": "social_security_number",
    "sécurité sociale": "social_security_number",
    "social_security_number": "social_security_number",
    # notes
    "notes": "notes",
    "commentaire": "notes",
    "commentaires": "notes",
}

_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
_PHONE_RE = re.compile(r"^[\d\s\+\-\.\(\)]{6,20}$")


def _detect_delimiter(sample: str) -> str:
    """Auto-detect CSV delimiter from a text sample."""
    for delim in [";", ",", "\t"]:
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample, delimiters=delim)
            return dialect.delimiter
        except csv.Error:
            continue
    # Count occurrences as fallback
    counts = {d: sample.count(d) for d in [";", ",", "\t"]}
    return max(counts, key=counts.get)  # type: ignore[arg-type]


def _normalize_column(name: str) -> str:
    """Normalize a column header for mapping lookup."""
    return name.strip().lower().replace("_", " ").replace("-", " ").strip()


def _parse_date(val: str) -> str | None:
    """Try to parse a date string into YYYY-MM-DD format."""
    from datetime import datetime as _dt

    val = val.strip()
    if not val:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%y"):
        try:
            return _dt.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _read_excel_rows(file_content: bytes) -> list[dict[str, str]]:
    """Read rows from an Excel (.xlsx) file and return list of dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [str(h).strip() if h else "" for h in header_row]
    result: list[dict[str, str]] = []
    for row in rows_iter:
        row_dict: dict[str, str] = {}
        for idx, cell_val in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = str(cell_val).strip() if cell_val is not None else ""
        result.append(row_dict)
    wb.close()
    return result


def import_from_file(
    db: Session,
    tenant_id: int,
    file_content: bytes,
    filename: str,
    user_id: int,
) -> ClientImportResult:
    """Import clients from CSV or Excel file with auto-detection and validation."""
    is_excel = filename.lower().endswith((".xlsx", ".xls"))

    # Parse rows
    if is_excel:
        raw_rows = _read_excel_rows(file_content)
    else:
        text = file_content.decode("utf-8-sig")
        delimiter = _detect_delimiter(text.split("\n")[0] if text else "")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        raw_rows = list(reader)

    if not raw_rows:
        return ClientImportResult(imported=0, updated=0, skipped=0, errors=[])

    # Build column mapping from file headers
    sample_keys = list(raw_rows[0].keys())
    col_mapping: dict[str, str] = {}
    for orig_key in sample_keys:
        normalized = _normalize_column(orig_key)
        if normalized in _COLUMN_MAP:
            col_mapping[orig_key] = _COLUMN_MAP[normalized]

    imported = 0
    updated = 0
    skipped = 0
    errors: list[ClientImportError] = []

    for i, raw_row in enumerate(raw_rows, start=2):
        try:
            # Map columns
            mapped: dict[str, str] = {}
            for orig_key, field_name in col_mapping.items():
                val = raw_row.get(orig_key, "").strip()
                if val:
                    mapped[field_name] = val

            last_name = mapped.get("last_name", "")
            first_name = mapped.get("first_name", "")
            if not last_name:
                skipped += 1
                errors.append(ClientImportError(line=i, reason="Nom de famille manquant"))
                continue

            # Validate email
            email = mapped.get("email")
            if email and not _EMAIL_RE.match(email):
                errors.append(ClientImportError(line=i, reason=f"Email invalide : {email}"))
                email = None

            # Validate phone
            phone = mapped.get("phone")
            if phone and not _PHONE_RE.match(phone):
                errors.append(ClientImportError(line=i, reason=f"Telephone invalide : {phone}"))
                phone = None

            # Parse birth_date
            birth_date_str = mapped.get("birth_date")
            birth_date = _parse_date(birth_date_str) if birth_date_str else None

            # Check for existing client (by email or name match) for update
            existing: Customer | None = None
            if email:
                existing = db.scalars(
                    select(Customer).where(
                        Customer.tenant_id == tenant_id,
                        Customer.email == email,
                        Customer.deleted_at.is_(None),
                    )
                ).first()

            if existing:
                # Update existing client with new non-empty fields
                changed = False
                for attr, val in [
                    ("first_name", first_name),
                    ("last_name", last_name),
                    ("phone", phone),
                    ("address", mapped.get("address")),
                    ("city", mapped.get("city")),
                    ("postal_code", mapped.get("postal_code")),
                    ("social_security_number", mapped.get("social_security_number")),
                    ("notes", mapped.get("notes")),
                ]:
                    if val and val != getattr(existing, attr, None):
                        setattr(existing, attr, val)
                        changed = True
                if birth_date:
                    from datetime import date as _date

                    bd = _date.fromisoformat(birth_date)
                    if bd != existing.birth_date:
                        existing.birth_date = bd
                        changed = True
                if changed:
                    updated += 1
                else:
                    skipped += 1
            else:
                # Create new client
                customer_kwargs: dict[str, object] = {
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "phone": phone,
                    "address": mapped.get("address"),
                    "city": mapped.get("city"),
                    "postal_code": mapped.get("postal_code"),
                    "social_security_number": mapped.get("social_security_number"),
                    "notes": mapped.get("notes"),
                }
                if birth_date:
                    from datetime import date as _date
                    customer_kwargs["birth_date"] = _date.fromisoformat(birth_date)
                customer = Customer(
                    tenant_id=tenant_id,
                    **{k: v for k, v in customer_kwargs.items() if v},
                )
                db.add(customer)
                imported += 1
        except Exception as e:
            errors.append(ClientImportError(line=i, reason=str(e)))
            skipped += 1

    db.commit()
    audit_service.log_action(
        db,
        tenant_id,
        user_id,
        "create",
        "client_import",
        0,
        new_value={"imported": imported, "updated": updated, "skipped": skipped},
    )
    logger.info(
        "file_import_completed",
        tenant_id=tenant_id,
        imported=imported,
        updated=updated,
        skipped=skipped,
        filename=filename,
    )
    return ClientImportResult(imported=imported, updated=updated, skipped=skipped, errors=errors[:50])


def generate_import_template() -> bytes:
    """Generate a CSV template for client import."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Nom", "Prenom", "Email", "Telephone", "Date de naissance",
        "Adresse", "Ville", "Code postal", "N° Secu", "Notes",
    ])
    writer.writerow([
        "Dupont", "Jean", "jean.dupont@email.fr", "06 12 34 56 78", "15/03/1985",
        "12 rue de la Paix", "Paris", "75001", "1850375012345", "Client fidele",
    ])
    return output.getvalue().encode("utf-8-sig")
