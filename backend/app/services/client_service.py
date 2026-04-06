import csv
import io
import re

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessError, NotFoundError
from app.core.logging import get_logger
from app.domain.schemas.clients import (
    ClientCreate,
    ClientImportError,
    ClientImportResult,
    ClientListResponse,
    ClientMergeResult,
    ClientResponse,
    ClientUpdate,
    DuplicateGroup,
)
from app.models.client import Customer
from app.repositories import client_repo
from app.services import audit_service

logger = get_logger("client_service")


def search_clients(
    db: Session,
    tenant_id: int,
    query: str,
    page: int,
    page_size: int,
    include_deleted: bool = False,
) -> ClientListResponse:
    items, total = client_repo.search(db, tenant_id, query, page, page_size, include_deleted=include_deleted)
    return ClientListResponse(
        items=[ClientResponse.model_validate(c) for c in items],
        total=total,
        page=page,
        page_size=page_size,
    )


def get_client(db: Session, tenant_id: int, client_id: int) -> ClientResponse:
    customer = client_repo.get_by_id_active(db, client_id=client_id, tenant_id=tenant_id)
    if not customer:
        raise NotFoundError("client", client_id)
    return ClientResponse.model_validate(customer)


def create_client(db: Session, tenant_id: int, payload: ClientCreate, user_id: int) -> ClientResponse:
    customer = client_repo.create(db, tenant_id=tenant_id, **payload.model_dump(exclude_none=True))
    audit_service.log_action(
        db,
        tenant_id,
        user_id,
        "create",
        "client",
        customer.id,
        new_value=payload.model_dump(exclude_none=True),
    )
    logger.info("client_created", tenant_id=tenant_id, client_id=customer.id, user_id=user_id)
    return ClientResponse.model_validate(customer)


def update_client(db: Session, tenant_id: int, client_id: int, payload: ClientUpdate, user_id: int) -> ClientResponse:
    customer = client_repo.get_by_id_active(db, client_id=client_id, tenant_id=tenant_id)
    if not customer:
        raise NotFoundError("client", client_id)
    updated = client_repo.update(db, customer, **payload.model_dump(exclude_unset=True))
    audit_service.log_action(
        db,
        tenant_id,
        user_id,
        "update",
        "client",
        client_id,
        new_value=payload.model_dump(exclude_unset=True),
    )
    logger.info("client_updated", tenant_id=tenant_id, client_id=client_id, user_id=user_id)
    return ClientResponse.model_validate(updated)


def delete_client(db: Session, tenant_id: int, client_id: int, user_id: int) -> None:
    customer = client_repo.get_by_id_active(db, client_id=client_id, tenant_id=tenant_id)
    if not customer:
        raise NotFoundError("client", client_id)
    client_repo.delete(db, customer)
    audit_service.log_action(db, tenant_id, user_id, "delete", "client", client_id)
    logger.info("client_deleted", tenant_id=tenant_id, client_id=client_id, user_id=user_id)


def restore_client(db: Session, tenant_id: int, client_id: int, user_id: int) -> ClientResponse:
    customer = client_repo.get_by_id(db, client_id=client_id, tenant_id=tenant_id)
    if not customer:
        raise NotFoundError("client", client_id)
    if customer.deleted_at is None:
        raise NotFoundError("client", client_id)

    # Check for duplicate email before restoring
    if customer.email:
        conflict = db.scalars(
            select(Customer).where(
                Customer.tenant_id == tenant_id,
                Customer.email == customer.email,
                Customer.deleted_at.is_(None),
                Customer.id != customer.id,
            )
        ).first()
        if conflict:
            raise BusinessError(
                "DUPLICATE_EMAIL",
                f"Un client actif avec l'email {customer.email} existe deja",
            )

    restored = client_repo.restore(db, customer)
    audit_service.log_action(db, tenant_id, user_id, "restore", "client", client_id)
    logger.info("client_restored", tenant_id=tenant_id, client_id=client_id, user_id=user_id)
    return ClientResponse.model_validate(restored)


# ---------------------------------------------------------------------------
# Column name mapping: French variants -> internal field names
# ---------------------------------------------------------------------------
_COLUMN_MAP: dict[str, str] = {
    # last_name
    "nom": "last_name",
    "nom de famille": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    # first_name
    "prenom": "first_name",
    "prénom": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    # email
    "email": "email",
    "e-mail": "email",
    "mail": "email",
    "adresse email": "email",
    "adresse e-mail": "email",
    # phone
    "telephone": "phone",
    "téléphone": "phone",
    "tel": "phone",
    "tél": "phone",
    "phone": "phone",
    "portable": "phone",
    "mobile": "phone",
    # birth_date
    "date de naissance": "birth_date",
    "date_naissance": "birth_date",
    "naissance": "birth_date",
    "birth_date": "birth_date",
    # address
    "adresse": "address",
    "address": "address",
    # city
    "ville": "city",
    "city": "city",
    # postal_code
    "code postal": "postal_code",
    "code_postal": "postal_code",
    "cp": "postal_code",
    "postal_code": "postal_code",
    # social_security_number
    "n° sécu": "social_security_number",
    "n° secu": "social_security_number",
    "numero secu": "social_security_number",
    "numéro sécu": "social_security_number",
    "nir": "social_security_number",
    "securite sociale": "social_security_number",
    "sécurité sociale": "social_security_number",
    "social_security_number": "social_security_number",
    # notes
    "notes": "notes",
    "commentaire": "notes",
    "commentaires": "notes",
}

_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
_PHONE_RE = re.compile(r"^[\d\s\+\-\.\(\)]{6,20}$")


def _detect_delimiter(sample: str) -> str:
    """Auto-detect CSV delimiter from a text sample."""
    for delim in [";", ",", "\t"]:
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample, delimiters=delim)
            return dialect.delimiter
        except csv.Error:
            continue
    # Count occurrences as fallback
    counts = {d: sample.count(d) for d in [";", ",", "\t"]}
    return max(counts, key=counts.get)  # type: ignore[arg-type]


def _normalize_column(name: str) -> str:
    """Normalize a column header for mapping lookup."""
    return name.strip().lower().replace("_", " ").replace("-", " ").strip()


def _parse_date(val: str) -> str | None:
    """Try to parse a date string into YYYY-MM-DD format."""
    import re as _re
    from datetime import datetime as _dt

    val = val.strip()
    if not val:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%y"):
        try:
            return _dt.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _read_excel_rows(file_content: bytes) -> list[dict[str, str]]:
    """Read rows from an Excel (.xlsx) file and return list of dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [str(h).strip() if h else "" for h in header_row]
    result: list[dict[str, str]] = []
    for row in rows_iter:
        row_dict: dict[str, str] = {}
        for idx, cell_val in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = str(cell_val).strip() if cell_val is not None else ""
        result.append(row_dict)
    wb.close()
    return result


def import_from_file(
    db: Session,
    tenant_id: int,
    file_content: bytes,
    filename: str,
    user_id: int,
) -> ClientImportResult:
    """Import clients from CSV or Excel file with auto-detection and validation."""
    is_excel = filename.lower().endswith((".xlsx", ".xls"))

    # Parse rows
    if is_excel:
        raw_rows = _read_excel_rows(file_content)
    else:
        text = file_content.decode("utf-8-sig")
        delimiter = _detect_delimiter(text.split("\n")[0] if text else "")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        raw_rows = list(reader)

    if not raw_rows:
        return ClientImportResult(imported=0, updated=0, skipped=0, errors=[])

    # Build column mapping from file headers
    sample_keys = list(raw_rows[0].keys())
    col_mapping: dict[str, str] = {}
    for orig_key in sample_keys:
        normalized = _normalize_column(orig_key)
        if normalized in _COLUMN_MAP:
            col_mapping[orig_key] = _COLUMN_MAP[normalized]

    imported = 0
    updated = 0
    skipped = 0
    errors: list[ClientImportError] = []

    for i, raw_row in enumerate(raw_rows, start=2):
        try:
            # Map columns
            mapped: dict[str, str] = {}
            for orig_key, field_name in col_mapping.items():
                val = raw_row.get(orig_key, "").strip()
                if val:
                    mapped[field_name] = val

            last_name = mapped.get("last_name", "")
            first_name = mapped.get("first_name", "")
            if not last_name:
                skipped += 1
                errors.append(ClientImportError(line=i, reason="Nom de famille manquant"))
                continue

            # Validate email
            email = mapped.get("email")
            if email and not _EMAIL_RE.match(email):
                errors.append(ClientImportError(line=i, reason=f"Email invalide : {email}"))
                email = None

            # Validate phone
            phone = mapped.get("phone")
            if phone and not _PHONE_RE.match(phone):
                errors.append(ClientImportError(line=i, reason=f"Telephone invalide : {phone}"))
                phone = None

            # Parse birth_date
            birth_date_str = mapped.get("birth_date")
            birth_date = _parse_date(birth_date_str) if birth_date_str else None

            # Check for existing client (by email or name match) for update
            existing: Customer | None = None
            if email:
                existing = db.scalars(
                    select(Customer).where(
                        Customer.tenant_id == tenant_id,
                        Customer.email == email,
                        Customer.deleted_at.is_(None),
                    )
                ).first()

            if existing:
                # Update existing client with new non-empty fields
                changed = False
                for attr, val in [
                    ("first_name", first_name),
                    ("last_name", last_name),
                    ("phone", phone),
                    ("address", mapped.get("address")),
                    ("city", mapped.get("city")),
                    ("postal_code", mapped.get("postal_code")),
                    ("social_security_number", mapped.get("social_security_number")),
                    ("notes", mapped.get("notes")),
                ]:
                    if val and val != getattr(existing, attr, None):
                        setattr(existing, attr, val)
                        changed = True
                if birth_date:
                    from datetime import date as _date

                    bd = _date.fromisoformat(birth_date)
                    if bd != existing.birth_date:
                        existing.birth_date = bd
                        changed = True
                if changed:
                    updated += 1
                else:
                    skipped += 1
            else:
                # Create new client
                customer_kwargs: dict[str, object] = {
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "phone": phone,
                    "address": mapped.get("address"),
                    "city": mapped.get("city"),
                    "postal_code": mapped.get("postal_code"),
                    "social_security_number": mapped.get("social_security_number"),
                    "notes": mapped.get("notes"),
                }
                if birth_date:
                    from datetime import date as _date
                    customer_kwargs["birth_date"] = _date.fromisoformat(birth_date)
                customer = Customer(
                    tenant_id=tenant_id,
                    **{k: v for k, v in customer_kwargs.items() if v},
                )
                db.add(customer)
                imported += 1
        except Exception as e:
            errors.append(ClientImportError(line=i, reason=str(e)))
            skipped += 1

    db.commit()
    audit_service.log_action(
        db,
        tenant_id,
        user_id,
        "create",
        "client_import",
        0,
        new_value={"imported": imported, "updated": updated, "skipped": skipped},
    )
    logger.info(
        "file_import_completed",
        tenant_id=tenant_id,
        imported=imported,
        updated=updated,
        skipped=skipped,
        filename=filename,
    )
    return ClientImportResult(imported=imported, updated=updated, skipped=skipped, errors=errors[:50])


def generate_import_template() -> bytes:
    """Generate a CSV template for client import."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Nom", "Prenom", "Email", "Telephone", "Date de naissance",
        "Adresse", "Ville", "Code postal", "N° Secu", "Notes",
    ])
    writer.writerow([
        "Dupont", "Jean", "jean.dupont@email.fr", "06 12 34 56 78", "15/03/1985",
        "12 rue de la Paix", "Paris", "75001", "1850375012345", "Client fidele",
    ])
    return output.getvalue().encode("utf-8-sig")


def find_duplicates(db: Session, tenant_id: int) -> list[DuplicateGroup]:
    """Find potential duplicate clients by case-insensitive name matching."""
    dupes = db.execute(
        select(
            func.lower(Customer.last_name),
            func.lower(Customer.first_name),
            func.count().label("cnt"),
        )
        .where(Customer.tenant_id == tenant_id, Customer.deleted_at.is_(None))
        .group_by(func.lower(Customer.last_name), func.lower(Customer.first_name))
        .having(func.count() > 1)
    ).all()
    results: list[DuplicateGroup] = []
    for last, first, cnt in dupes:
        clients = db.scalars(
            select(Customer).where(
                Customer.tenant_id == tenant_id,
                func.lower(Customer.last_name) == last,
                func.lower(Customer.first_name) == first,
                Customer.deleted_at.is_(None),
            )
        ).all()
        results.append(
            DuplicateGroup(
                name=f"{first} {last}",
                count=cnt,
                clients=[ClientResponse.model_validate(c) for c in clients],
            )
        )
    return results


def upload_avatar(
    db: Session, tenant_id: int, client_id: int, file_data: bytes, content_type: str, user_id: int
) -> str:
    """Upload a client avatar to MinIO and store the URL."""
    import uuid

    from app.core.config import settings
    from app.integrations.storage import storage

    customer = client_repo.get_by_id_active(db, client_id=client_id, tenant_id=tenant_id)
    if not customer:
        raise NotFoundError("client", client_id)

    ext = "jpg" if "jpeg" in content_type or "jpg" in content_type else "png"
    storage_key = f"tenants/{tenant_id}/avatars/{client_id}/{uuid.uuid4().hex}.{ext}"

    storage.upload_file(
        bucket=settings.s3_bucket,
        key=storage_key,
        file_data=file_data,
        content_type=content_type,
    )

    avatar_url = f"/api/v1/clients/{client_id}/avatar"
    customer.avatar_url = storage_key
    db.commit()
    db.refresh(customer)

    audit_service.log_action(
        db,
        tenant_id,
        user_id,
        "update",
        "client_avatar",
        client_id,
        new_value={"avatar_key": storage_key},
    )
    logger.info("avatar_uploaded", tenant_id=tenant_id, client_id=client_id)
    return avatar_url


def get_avatar_url(db: Session, tenant_id: int, client_id: int) -> str:
    """Get the MinIO presigned URL for a client avatar."""
    from app.core.config import settings
    from app.integrations.storage import storage

    customer = client_repo.get_by_id_active(db, client_id=client_id, tenant_id=tenant_id)
    if not customer or not customer.avatar_url:
        raise NotFoundError("avatar", client_id)

    url = storage.get_download_url(
        bucket=settings.s3_bucket,
        key=customer.avatar_url,
        expires=3600,
    )
    return url


def merge_clients(
    db: Session, tenant_id: int, keep_id: int, merge_id: int, user_id: int
) -> ClientMergeResult:
    """Merge merge_id into keep_id. Transfer all related data, then soft-delete merge_id."""
    from app.models.case import Case
    from app.models.client_mutuelle import ClientMutuelle
    from app.models.cosium_data import (
        CosiumDocument,
        CosiumInvoice,
        CosiumPayment,
        CosiumPrescription,
    )
    from app.models.interaction import Interaction
    from app.models.marketing import MarketingConsent, MessageLog, SegmentMembership
    from app.models.pec import PecRequest
    from app.models.pec_preparation import PecPreparation

    if keep_id == merge_id:
        raise BusinessError("Impossible de fusionner un client avec lui-meme", "MERGE_SAME_CLIENT")

    keep_client = client_repo.get_by_id_active(db, client_id=keep_id, tenant_id=tenant_id)
    if not keep_client:
        raise NotFoundError("client", keep_id)

    merge_client = client_repo.get_by_id_active(db, client_id=merge_id, tenant_id=tenant_id)
    if not merge_client:
        raise NotFoundError("client", merge_id)

    # Fill empty fields on keep_client from merge_client
    fillable_fields = [
        "phone", "email", "birth_date", "address", "street_number",
        "street_name", "city", "postal_code", "social_security_number",
        "optician_name", "ophthalmologist_id", "notes", "cosium_id",
        "customer_number", "mobile_phone_country", "site_id",
    ]
    fields_filled: list[str] = []
    for field in fillable_fields:
        keep_val = getattr(keep_client, field, None)
        merge_val = getattr(merge_client, field, None)
        if not keep_val and merge_val:
            setattr(keep_client, field, merge_val)
            fields_filled.append(field)

    # Transfer cases
    cases_transferred = db.execute(
        select(func.count()).select_from(Case).where(
            Case.customer_id == merge_id, Case.tenant_id == tenant_id
        )
    ).scalar_one()
    if cases_transferred:
        db.execute(
            Case.__table__.update()
            .where(Case.customer_id == merge_id, Case.tenant_id == tenant_id)
            .values(customer_id=keep_id)
        )

    # Transfer interactions
    interactions_transferred = db.execute(
        select(func.count()).select_from(Interaction).where(
            Interaction.client_id == merge_id, Interaction.tenant_id == tenant_id
        )
    ).scalar_one()
    if interactions_transferred:
        db.execute(
            Interaction.__table__.update()
            .where(Interaction.client_id == merge_id, Interaction.tenant_id == tenant_id)
            .values(client_id=keep_id)
        )

    # Transfer PEC requests (PecRequest links via case_id, which is transferred above,
    # so we count them for reporting but they follow the cases automatically)
    pec_transferred = db.execute(
        select(func.count()).select_from(PecRequest).where(
            PecRequest.case_id.in_(
                select(Case.id).where(Case.customer_id == keep_id, Case.tenant_id == tenant_id)
            ),
            PecRequest.tenant_id == tenant_id,
        )
    ).scalar_one() if cases_transferred else 0

    # Transfer PEC preparations
    db.execute(
        PecPreparation.__table__.update()
        .where(PecPreparation.customer_id == merge_id, PecPreparation.tenant_id == tenant_id)
        .values(customer_id=keep_id)
    )

    # Transfer client mutuelles
    db.execute(
        ClientMutuelle.__table__.update()
        .where(ClientMutuelle.customer_id == merge_id, ClientMutuelle.tenant_id == tenant_id)
        .values(customer_id=keep_id)
    )

    # Transfer marketing data
    marketing_transferred = 0
    for model in [MarketingConsent, SegmentMembership, MessageLog]:
        cnt = db.execute(
            select(func.count()).select_from(model).where(
                model.client_id == merge_id, model.tenant_id == tenant_id
            )
        ).scalar_one()
        marketing_transferred += cnt
        if cnt:
            db.execute(
                model.__table__.update()
                .where(model.client_id == merge_id, model.tenant_id == tenant_id)
                .values(client_id=keep_id)
            )

    # Transfer cosium data references
    cosium_transferred = 0
    for cosium_model in [CosiumInvoice, CosiumPayment, CosiumDocument, CosiumPrescription]:
        if hasattr(cosium_model, "customer_id") and hasattr(cosium_model, "tenant_id"):
            cnt = db.execute(
                select(func.count()).select_from(cosium_model).where(
                    cosium_model.customer_id == merge_id, cosium_model.tenant_id == tenant_id
                )
            ).scalar_one()
            cosium_transferred += cnt
            if cnt:
                db.execute(
                    cosium_model.__table__.update()
                    .where(cosium_model.customer_id == merge_id, cosium_model.tenant_id == tenant_id)
                    .values(customer_id=keep_id)
                )

    # Soft-delete merged client
    client_repo.delete(db, merge_client)

    db.commit()
    db.refresh(keep_client)

    audit_service.log_action(
        db,
        tenant_id,
        user_id,
        "merge",
        "client",
        keep_id,
        new_value={
            "merged_from": merge_id,
            "cases_transferred": cases_transferred,
            "interactions_transferred": interactions_transferred,
            "pec_transferred": pec_transferred,
            "marketing_transferred": marketing_transferred,
            "fields_filled": fields_filled,
        },
    )
    logger.info(
        "clients_merged",
        tenant_id=tenant_id,
        keep_id=keep_id,
        merge_id=merge_id,
        user_id=user_id,
    )

    return ClientMergeResult(
        kept_client=ClientResponse.model_validate(keep_client),
        cases_transferred=cases_transferred,
        interactions_transferred=interactions_transferred,
        pec_transferred=pec_transferred,
        marketing_transferred=marketing_transferred,
        cosium_data_transferred=cosium_transferred,
        fields_filled=fields_filled,
        merged_client_deleted=True,
    )
