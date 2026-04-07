"""Service d'export des operations batch (Excel/CSV)."""

import io
from datetime import datetime

from fastapi.responses import StreamingResponse


def export_batch_excel(enriched: dict) -> StreamingResponse:
    """Generate a professional 2-sheet Excel export of batch summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    bold_font = Font(bold=True)
    green_fill = PatternFill(
        start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
    )
    red_fill = PatternFill(
        start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
    )
    amber_fill = PatternFill(
        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
    )

    batch = enriched["batch"]
    items = enriched["items"]

    # Sheet 1: Resume
    ws1 = wb.active
    ws1.title = "Resume"

    info_rows = [
        ("Operation batch", f"#{batch.id}"),
        ("Code marketing", batch.marketing_code),
        ("Label", batch.label or "-"),
        ("Statut", batch.status),
        ("Date debut", batch.started_at.strftime("%d/%m/%Y %H:%M") if batch.started_at else "-"),
        ("Date fin", batch.completed_at.strftime("%d/%m/%Y %H:%M") if batch.completed_at else "-"),
        ("", ""),
        ("Total clients", batch.total_clients),
        ("Clients prets", batch.clients_prets),
        ("Clients incomplets", batch.clients_incomplets),
        ("Clients en conflit", batch.clients_en_conflit),
        ("Clients en erreur", batch.clients_erreur),
    ]
    for row_idx, (label, value) in enumerate(info_rows, 1):
        c_label = ws1.cell(row=row_idx, column=1, value=label)
        c_label.font = bold_font
        c_label.border = thin_border
        c_val = ws1.cell(row=row_idx, column=2, value=value)
        c_val.border = thin_border
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40

    # Sheet 2: Detail clients
    ws2 = wb.create_sheet("Detail clients")
    headers = [
        "Client", "Telephone", "Email", "N. Secu", "Mutuelle",
        "Score (%)", "Statut", "Erreurs", "Alertes", "PEC ID",
        "Message erreur", "Traite le",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_fills = {
        "pret": green_fill,
        "incomplet": amber_fill,
        "conflit": red_fill,
        "erreur": red_fill,
    }

    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.get("customer_name", ""),
            item.get("phone", ""),
            item.get("email", ""),
            item.get("social_security_number", ""),
            item.get("mutuelle_name", ""),
            round(item.get("completude_score", 0), 1),
            item.get("status", ""),
            item.get("errors_count", 0),
            item.get("warnings_count", 0),
            item.get("pec_preparation_id") or "",
            item.get("error_message") or "",
            item.get("processed_at", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
        status_cell = ws2.cell(row=row_idx, column=7)
        fill = status_fills.get(item.get("status"))
        if fill:
            status_cell.fill = fill

    widths = [30, 18, 30, 18, 25, 10, 14, 10, 10, 10, 40, 20]
    for i, w in enumerate(widths):
        ws2.column_dimensions[chr(65 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"batch_{batch.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_batch_csv(enriched: dict) -> StreamingResponse:
    """Fallback CSV export."""
    import csv

    batch = enriched["batch"]
    items = enriched["items"]

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([
        "Client", "Telephone", "Email", "N. Secu", "Mutuelle",
        "Score", "Statut", "Erreurs", "Alertes", "PEC ID", "Erreur",
    ])
    for item in items:
        writer.writerow([
            item.get("customer_name", ""),
            item.get("phone", ""),
            item.get("email", ""),
            item.get("social_security_number", ""),
            item.get("mutuelle_name", ""),
            round(item.get("completude_score", 0), 1),
            item.get("status", ""),
            item.get("errors_count", 0),
            item.get("warnings_count", 0),
            item.get("pec_preparation_id") or "",
            item.get("error_message") or "",
        ])

    content = buf.getvalue().encode("utf-8-sig")
    filename = f"batch_{batch.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
