"""Service d'export CSV et Excel generique, plus balance clients XLSX.

Les exports FEC et PDF sont dans export_fec.py et export_pdf.py.
Re-exports fournis pour compatibilite ascendante.
"""

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger, log_operation
from app.models import (
    AuditLog,
    Campaign,
    Customer,
    Devis,
    Facture,
    Payment,
    PecRequest,
    Reminder,
)

logger = get_logger("export_service")

ENTITY_CONFIGS: dict[str, dict] = {
    "clients": {
        "model": Customer,
        "columns": ["id", "first_name", "last_name", "email", "phone", "city", "postal_code", "created_at"],
        "headers": ["ID", "Prenom", "Nom", "Email", "Telephone", "Ville", "Code postal", "Date creation"],
    },
    "factures": {
        "model": Facture,
        "columns": ["id", "numero", "montant_ht", "tva", "montant_ttc", "status", "date_emission", "created_at"],
        "headers": ["ID", "Numero", "Montant HT", "TVA", "Montant TTC", "Statut", "Date emission", "Date creation"],
    },
    "paiements": {
        "model": Payment,
        "columns": [
            "id",
            "case_id",
            "payer_type",
            "mode_paiement",
            "amount_due",
            "amount_paid",
            "status",
            "created_at",
        ],
        "headers": ["ID", "Dossier", "Type payeur", "Mode", "Montant du", "Montant paye", "Statut", "Date"],
    },
    "devis": {
        "model": Devis,
        "columns": [
            "id",
            "numero",
            "status",
            "montant_ht",
            "tva",
            "montant_ttc",
            "part_secu",
            "part_mutuelle",
            "reste_a_charge",
            "created_at",
        ],
        "headers": ["ID", "Numero", "Statut", "HT", "TVA", "TTC", "Part secu", "Part mutuelle", "RAC", "Date"],
    },
    "pec": {
        "model": PecRequest,
        "columns": ["id", "case_id", "organization_id", "montant_demande", "montant_accorde", "status", "created_at"],
        "headers": ["ID", "Dossier", "Organisme", "Montant demande", "Montant accorde", "Statut", "Date"],
    },
    "relances": {
        "model": Reminder,
        "columns": ["id", "target_type", "target_id", "channel", "status", "content", "created_at"],
        "headers": ["ID", "Type cible", "ID cible", "Canal", "Statut", "Contenu", "Date"],
    },
    "campagnes": {
        "model": Campaign,
        "columns": ["id", "name", "channel", "status", "sent_at", "created_at"],
        "headers": ["ID", "Nom", "Canal", "Statut", "Date envoi", "Date creation"],
    },
    "audit_logs": {
        "model": AuditLog,
        "columns": ["id", "user_id", "action", "entity_type", "entity_id", "created_at"],
        "headers": ["ID", "Utilisateur", "Action", "Type entite", "ID entite", "Date"],
    },
}


def _get_rows(
    db: Session, tenant_id: int, entity_type: str, date_from: datetime | None = None, date_to: datetime | None = None
) -> list[list]:
    config = ENTITY_CONFIGS.get(entity_type)
    if not config:
        return []

    model = config["model"]
    columns = config["columns"]

    max_export_rows = 50000

    q = select(model)
    if hasattr(model, "tenant_id"):
        q = q.where(model.tenant_id == tenant_id)
    if date_from and hasattr(model, "created_at"):
        q = q.where(model.created_at >= date_from)
    if date_to and hasattr(model, "created_at"):
        q = q.where(model.created_at <= date_to)

    items = db.scalars(q.order_by(model.id.desc()).limit(max_export_rows)).all()
    rows = []
    for item in items:
        row = []
        for col in columns:
            val = getattr(item, col, "")
            if isinstance(val, datetime):
                val = val.strftime("%d/%m/%Y %H:%M")
            elif val is None:
                val = ""
            row.append(val)
        rows.append(row)
    return rows


@log_operation("export_csv")
def export_to_csv(
    db: Session, tenant_id: int, entity_type: str, date_from: datetime | None = None, date_to: datetime | None = None
) -> bytes:
    config = ENTITY_CONFIGS.get(entity_type)
    if not config:
        return b""

    rows = _get_rows(db, tenant_id, entity_type, date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(config["headers"])
    writer.writerows(rows)

    logger.info("csv_exported", tenant_id=tenant_id, entity_type=entity_type, rows=len(rows))
    return output.getvalue().encode("utf-8-sig")


@log_operation("export_xlsx")
def export_to_xlsx(
    db: Session, tenant_id: int, entity_type: str, date_from: datetime | None = None, date_to: datetime | None = None
) -> bytes:
    config = ENTITY_CONFIGS.get(entity_type)
    if not config:
        return b""

    rows = _get_rows(db, tenant_id, entity_type, date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = entity_type.capitalize()
    ws.append(config["headers"])
    for row in rows:
        ws.append([str(v) for v in row])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("xlsx_exported", tenant_id=tenant_id, entity_type=entity_type, rows=len(rows))
    return output.getvalue()


@log_operation("export_balance_clients_xlsx")
def export_balance_clients_xlsx(
    db: Session,
    tenant_id: int,
    date_from=None,
    date_to=None,
) -> bytes:
    """Generate an Excel balance report of outstanding client debts."""
    from app.services.export_pdf import _get_balance_rows

    rows = _get_balance_rows(db, tenant_id, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Clients"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Client", "Nb Factures", "Total Facture (EUR)", "Total Impaye (EUR)", "Derniere Facture"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    total_facture = 0.0
    total_impaye = 0.0
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["client"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=row["nb_factures"]).border = thin_border
        c_fact = ws.cell(row=row_idx, column=3, value=row["total_facture"])
        c_fact.number_format = "#,##0.00"
        c_fact.border = thin_border
        c_imp = ws.cell(row=row_idx, column=4, value=row["total_impaye"])
        c_imp.number_format = "#,##0.00"
        c_imp.border = thin_border
        date_val = row["derniere_facture"].strftime("%d/%m/%Y") if row["derniere_facture"] else ""
        ws.cell(row=row_idx, column=5, value=date_val).border = thin_border
        total_facture += row["total_facture"]
        total_impaye += row["total_impaye"]

    # Total row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(rows)).font = Font(bold=True)
    c_tf = ws.cell(row=total_row, column=3, value=round(total_facture, 2))
    c_tf.number_format = "#,##0.00"
    c_tf.font = Font(bold=True)
    c_ti = ws.cell(row=total_row, column=4, value=round(total_impaye, 2))
    c_ti.number_format = "#,##0.00"
    c_ti.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("balance_clients_xlsx_exported", tenant_id=tenant_id, rows=len(rows))
    return output.getvalue()


@log_operation("export_clients_complet_xlsx")
def export_clients_complet_xlsx(
    db: Session,
    tenant_id: int,
    date_from=None,
    date_to=None,
    has_email: bool | None = None,
    has_cosium_id: bool | None = None,
) -> bytes:
    """Export ALL clients with ALL data columns in styled Excel format."""
    from datetime import datetime as _dt

    q = select(Customer).where(Customer.tenant_id == tenant_id, Customer.deleted_at.is_(None))
    if date_from:
        q = q.where(Customer.created_at >= _dt.combine(date_from, _dt.min.time()))
    if date_to:
        q = q.where(Customer.created_at <= _dt.combine(date_to, _dt.max.time()))
    if has_email is True:
        q = q.where(Customer.email.isnot(None), Customer.email != "")
    elif has_email is False:
        q = q.where((Customer.email.is_(None)) | (Customer.email == ""))
    if has_cosium_id is True:
        q = q.where(Customer.cosium_id.isnot(None), Customer.cosium_id != "")
    elif has_cosium_id is False:
        q = q.where((Customer.cosium_id.is_(None)) | (Customer.cosium_id == ""))

    clients = db.scalars(q.order_by(Customer.last_name, Customer.first_name).limit(50000)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients Complet"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Nom", "Prenom", "Date de naissance", "Telephone", "Email",
        "Adresse", "Ville", "Code postal", "N° Secu",
        "N° Client Cosium", "N° Client", "Opticien",
        "Notes", "Date creation",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, c in enumerate(clients, 2):
        vals = [
            c.last_name,
            c.first_name,
            c.birth_date.strftime("%d/%m/%Y") if c.birth_date else "",
            c.phone or "",
            c.email or "",
            c.address or "",
            c.city or "",
            c.postal_code or "",
            c.social_security_number or "",
            c.cosium_id or "",
            c.customer_number or "",
            c.optician_name or "",
            c.notes or "",
            c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "",
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Column widths
    widths = [20, 15, 16, 18, 30, 35, 20, 12, 18, 18, 15, 20, 30, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("clients_complet_xlsx_exported", tenant_id=tenant_id, rows=len(clients))
    return output.getvalue()


@log_operation("export_pec_preparations_xlsx")
def export_pec_preparations_xlsx(
    db: Session,
    tenant_id: int,
    status: str | None = None,
) -> bytes:
    """Export PEC preparations as an Excel file.

    Columns: Client, Mutuelle, N Adherent, Correction OD/OG,
    Montant TTC, Part Secu, Part Mutuelle, RAC, Score, Documents.
    """
    import json

    from app.models.client import Customer as CustomerModel
    from app.models.pec_preparation import PecPreparation

    stmt = select(PecPreparation).where(PecPreparation.tenant_id == tenant_id)
    if status:
        stmt = stmt.where(PecPreparation.status == status)
    stmt = stmt.order_by(PecPreparation.created_at.desc())
    preps = list(db.scalars(stmt).all())

    # Fetch customer names
    customer_ids = list({p.customer_id for p in preps})
    customers_map: dict[int, str] = {}
    if customer_ids:
        rows = db.execute(
            select(CustomerModel.id, CustomerModel.first_name, CustomerModel.last_name).where(
                CustomerModel.id.in_(customer_ids),
            )
        ).all()
        for cid, fn, ln in rows:
            customers_map[cid] = f"{fn or ''} {ln or ''}".strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Preparations PEC"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Client",
        "Mutuelle",
        "N. Adherent",
        "OD (Sph/Cyl/Axe/Add)",
        "OG (Sph/Cyl/Axe/Add)",
        "Montant TTC",
        "Part Secu",
        "Part Mutuelle",
        "RAC",
        "Score (%)",
        "Statut",
        "Documents",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    def _get_pec_field(profile: dict, key: str) -> str:
        field = profile.get(key)
        if isinstance(field, dict):
            return str(field.get("value", "-"))
        return str(field) if field else "-"

    def _correction_str(profile: dict, prefix: str) -> str:
        sph = _get_pec_field(profile, f"sphere_{prefix}")
        cyl = _get_pec_field(profile, f"cylinder_{prefix}")
        axe = _get_pec_field(profile, f"axis_{prefix}")
        add = _get_pec_field(profile, f"addition_{prefix}")
        return f"{sph} / {cyl} / {axe} / {add}"

    for row_idx, prep in enumerate(preps, 2):
        profile: dict = {}
        if prep.consolidated_data:
            try:
                profile = json.loads(prep.consolidated_data)
            except (json.JSONDecodeError, TypeError):
                pass

        customer_name = customers_map.get(prep.customer_id, f"Client #{prep.customer_id}")
        mutuelle = _get_pec_field(profile, "mutuelle_nom")
        num_adherent = _get_pec_field(profile, "mutuelle_numero_adherent")
        od_str = _correction_str(profile, "od")
        og_str = _correction_str(profile, "og")

        montant_ttc = _get_pec_field(profile, "montant_ttc")
        part_secu = _get_pec_field(profile, "part_secu")
        part_mutuelle = _get_pec_field(profile, "part_mutuelle")
        rac = _get_pec_field(profile, "reste_a_charge")
        score = prep.completude_score

        doc_count = len(prep.documents) if prep.documents else 0

        row_data = [
            customer_name,
            mutuelle,
            num_adherent,
            od_str,
            og_str,
            montant_ttc,
            part_secu,
            part_mutuelle,
            rac,
            round(score, 1),
            prep.status,
            f"{doc_count} doc(s)",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Column widths
    widths = [30, 25, 18, 25, 25, 14, 14, 14, 14, 10, 15, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("pec_preparations_xlsx_exported", tenant_id=tenant_id, count=len(preps))
    return output.getvalue()


# ---------------------------------------------------------------------------
# Re-exports for backward compatibility.
# Callers that do `export_service.generate_fec(...)` or
# `export_service.export_dashboard_pdf(...)` still work.
# ---------------------------------------------------------------------------
from app.services.export_fec import generate_fec  # noqa: E402, F401
from app.services.export_pdf import (  # noqa: E402, F401
    export_balance_clients_pdf,
    export_dashboard_pdf,
)
