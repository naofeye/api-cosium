"""Unit tests for export_service.

Tests the CSV and XLSX export functions directly at the service layer,
using the in-memory SQLite db fixture.  PDF/XLSX specialist generators
are mocked where needed so the tests stay fast and isolated.
"""

import csv
import io
from datetime import datetime
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from app.models.client import Customer
from app.models.tenant import Tenant
from app.services import export_service


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _seed_customers(db, tenant_id: int, count: int = 3) -> list[Customer]:
    customers = []
    for i in range(count):
        c = Customer(
            tenant_id=tenant_id,
            first_name=f"Prenom{i}",
            last_name=f"Nom{i}",
            email=f"client{i}@test.com",
            phone=f"06000000{i:02d}",
            city="Paris",
            postal_code="75001",
        )
        db.add(c)
        db.flush()
        customers.append(c)
    db.commit()
    return customers


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

class TestExportToCsv:
    def test_csv_returns_bytes_with_bom(self, db, default_tenant: Tenant):
        """CSV output is bytes and starts with UTF-8 BOM."""
        _seed_customers(db, default_tenant.id, count=2)
        result = export_service.export_to_csv(db, default_tenant.id, "clients")
        assert isinstance(result, bytes)
        # UTF-8-sig BOM marker
        assert result[:3] == b"\xef\xbb\xbf"

    def test_csv_contains_correct_headers(self, db, default_tenant: Tenant):
        """CSV first row matches configured headers for 'clients'."""
        result = export_service.export_to_csv(db, default_tenant.id, "clients")
        text = result.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), delimiter=";")
        headers = next(reader)
        assert "Prenom" in headers
        assert "Nom" in headers
        assert "Email" in headers

    def test_csv_contains_seeded_data(self, db, default_tenant: Tenant):
        """CSV rows include the seeded customer data."""
        _seed_customers(db, default_tenant.id, count=2)
        result = export_service.export_to_csv(db, default_tenant.id, "clients")
        text = result.decode("utf-8-sig")
        assert "Prenom0" in text
        assert "Nom1" in text

    def test_csv_tenant_isolation(self, db, default_tenant: Tenant):
        """Customers belonging to another tenant are excluded from the export."""
        from app.models import Organization, Tenant as TenantModel

        other_org = Organization(name="Autre Org", slug="autre-org", plan="solo")
        db.add(other_org)
        db.flush()
        other_tenant = TenantModel(
            organization_id=other_org.id,
            name="Autre Magasin",
            slug="autre-magasin",
        )
        db.add(other_tenant)
        db.flush()

        other_customer = Customer(
            tenant_id=other_tenant.id,
            first_name="Etranger",
            last_name="Invisible",
            email="autre@test.com",
        )
        db.add(other_customer)
        db.commit()

        result = export_service.export_to_csv(db, default_tenant.id, "clients")
        text = result.decode("utf-8-sig")
        assert "Invisible" not in text

    def test_csv_unknown_entity_returns_empty(self, db, default_tenant: Tenant):
        """Unknown entity type returns empty bytes."""
        result = export_service.export_to_csv(db, default_tenant.id, "NONEXISTENT")
        assert result == b""

    def test_csv_uses_semicolon_delimiter(self, db, default_tenant: Tenant):
        """Columns are delimited by semicolons, not commas."""
        result = export_service.export_to_csv(db, default_tenant.id, "clients")
        text = result.decode("utf-8-sig")
        first_line = text.splitlines()[0]
        assert ";" in first_line

    def test_csv_with_date_filter_from(self, db, default_tenant: Tenant):
        """date_from parameter is accepted without error."""
        _seed_customers(db, default_tenant.id)
        result = export_service.export_to_csv(
            db, default_tenant.id, "clients",
            date_from=datetime(2000, 1, 1),
        )
        assert isinstance(result, bytes)

    def test_csv_empty_table_returns_headers_only(self, db, default_tenant: Tenant):
        """When there are no rows the CSV still has the header line."""
        result = export_service.export_to_csv(db, default_tenant.id, "clients")
        text = result.decode("utf-8-sig")
        lines = [l for l in text.splitlines() if l.strip()]
        # Only the header row (no data rows)
        assert len(lines) == 1
        assert "Prenom" in lines[0]


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

class TestExportToXlsx:
    def test_xlsx_returns_bytes(self, db, default_tenant: Tenant):
        """XLSX output is non-empty bytes."""
        result = export_service.export_to_xlsx(db, default_tenant.id, "clients")
        assert isinstance(result, bytes)
        assert len(result) > 100

    def test_xlsx_is_valid_workbook(self, db, default_tenant: Tenant):
        """Returned bytes can be loaded as a valid openpyxl Workbook."""
        result = export_service.export_to_xlsx(db, default_tenant.id, "clients")
        wb = load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_xlsx_has_correct_sheet_name(self, db, default_tenant: Tenant):
        """Sheet name is the capitalised entity type."""
        result = export_service.export_to_xlsx(db, default_tenant.id, "clients")
        wb = load_workbook(io.BytesIO(result))
        assert "Clients" in wb.sheetnames

    def test_xlsx_header_row_matches_config(self, db, default_tenant: Tenant):
        """First row of the worksheet matches the configured headers."""
        result = export_service.export_to_xlsx(db, default_tenant.id, "clients")
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Prenom" in headers
        assert "Nom" in headers

    def test_xlsx_contains_seeded_data(self, db, default_tenant: Tenant):
        """Data rows contain the seeded customer values."""
        _seed_customers(db, default_tenant.id, count=2)
        result = export_service.export_to_xlsx(db, default_tenant.id, "clients")
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        all_values = [str(cell.value) for row in ws.iter_rows(min_row=2) for cell in row]
        assert any("Prenom0" in v for v in all_values)

    def test_xlsx_unknown_entity_returns_empty(self, db, default_tenant: Tenant):
        """Unknown entity type returns empty bytes."""
        result = export_service.export_to_xlsx(db, default_tenant.id, "NONEXISTENT")
        assert result == b""

    def test_xlsx_all_supported_entity_types(self, db, default_tenant: Tenant):
        """All configured entity types produce non-empty XLSX without error."""
        entity_types = list(export_service.ENTITY_CONFIGS.keys())
        for entity_type in entity_types:
            result = export_service.export_to_xlsx(db, default_tenant.id, entity_type)
            assert isinstance(result, bytes), f"XLSX for '{entity_type}' is not bytes"
            assert len(result) > 0, f"XLSX for '{entity_type}' is empty"


# ---------------------------------------------------------------------------
# Re-export surface — verify public API names are present
# ---------------------------------------------------------------------------

class TestReExportSurface:
    """export_service must expose all specialist functions via its public namespace.

    These tests validate that the module-level re-exports are in place so that
    callers using ``from app.services import export_service`` then
    ``export_service.export_balance_clients_pdf(...)`` continue to work.
    We don't call the real implementations here (they require reportlab/openpyxl
    pipelines with live data); we just assert the names are present and callable.
    """

    def test_pdf_balance_clients_is_exported(self):
        assert callable(getattr(export_service, "export_balance_clients_pdf", None))

    def test_pdf_dashboard_is_exported(self):
        assert callable(getattr(export_service, "export_dashboard_pdf", None))

    def test_pdf_monthly_report_is_exported(self):
        assert callable(getattr(export_service, "export_monthly_report_pdf", None))

    def test_xlsx_balance_clients_is_exported(self):
        assert callable(getattr(export_service, "export_balance_clients_xlsx", None))

    def test_xlsx_clients_complet_is_exported(self):
        assert callable(getattr(export_service, "export_clients_complet_xlsx", None))

    def test_xlsx_pec_preparations_is_exported(self):
        assert callable(getattr(export_service, "export_pec_preparations_xlsx", None))

    def test_fec_generate_is_exported(self):
        assert callable(getattr(export_service, "generate_fec", None))


# ---------------------------------------------------------------------------
# PDF export — orchestration with mocked specialist modules
# ---------------------------------------------------------------------------

class TestPdfExportOrchestration:
    @patch("app.services.export_service.export_balance_clients_pdf")
    def test_balance_clients_pdf_dispatched(self, mock_pdf, db, default_tenant: Tenant):
        """export_balance_clients_pdf can be patched in the export_service namespace."""
        mock_pdf.return_value = b"%PDF-balance"
        result = export_service.export_balance_clients_pdf(db, default_tenant.id)
        assert result == b"%PDF-balance"
        mock_pdf.assert_called_once_with(db, default_tenant.id)

    @patch("app.services.export_service.export_dashboard_pdf")
    def test_dashboard_pdf_dispatched(self, mock_pdf, db, default_tenant: Tenant):
        """export_dashboard_pdf can be patched in the export_service namespace."""
        mock_pdf.return_value = b"%PDF-dashboard"
        result = export_service.export_dashboard_pdf(db, default_tenant.id)
        assert result == b"%PDF-dashboard"

    @patch("app.services.export_service.export_monthly_report_pdf")
    def test_monthly_report_pdf_dispatched(self, mock_pdf, db, default_tenant: Tenant):
        """export_monthly_report_pdf can be patched in the export_service namespace."""
        mock_pdf.return_value = b"%PDF-monthly"
        result = export_service.export_monthly_report_pdf(db, default_tenant.id, year=2026, month=4)
        assert result == b"%PDF-monthly"


# ---------------------------------------------------------------------------
# XLSX specialist re-exports — patched at the export_service namespace
# ---------------------------------------------------------------------------

class TestXlsxSpecialistReexports:
    @patch("app.services.export_service.export_balance_clients_xlsx")
    def test_balance_xlsx_re_export(self, mock_xlsx, db, default_tenant: Tenant):
        """export_balance_clients_xlsx is patchable via the export_service namespace."""
        mock_xlsx.return_value = b"PK-xlsx-balance"
        result = export_service.export_balance_clients_xlsx(db, default_tenant.id)
        assert result == b"PK-xlsx-balance"

    @patch("app.services.export_service.export_clients_complet_xlsx")
    def test_clients_complet_xlsx_re_export(self, mock_xlsx, db, default_tenant: Tenant):
        """export_clients_complet_xlsx is patchable via the export_service namespace."""
        mock_xlsx.return_value = b"PK-xlsx-clients"
        result = export_service.export_clients_complet_xlsx(db, default_tenant.id)
        assert result == b"PK-xlsx-clients"


# ---------------------------------------------------------------------------
# _get_rows helper (indirect via export functions)
# ---------------------------------------------------------------------------

class TestGetRows:
    def test_date_values_formatted_as_french_datetime(self, db, default_tenant: Tenant):
        """datetime column values are formatted dd/mm/YYYY HH:MM in CSV output."""
        _seed_customers(db, default_tenant.id, count=1)
        result = export_service.export_to_csv(db, default_tenant.id, "clients")
        text = result.decode("utf-8-sig")
        # created_at is the last column — check format hint (day/month/year)
        import re
        assert re.search(r"\d{2}/\d{2}/\d{4}", text), "No French date format found in CSV"

    def test_max_export_rows_constant_is_positive(self):
        """Sanity: _MAX_EXPORT_ROWS is a positive integer."""
        assert export_service._MAX_EXPORT_ROWS > 0
