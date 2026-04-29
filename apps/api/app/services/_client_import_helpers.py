"""Helpers internes pour client_import_service : column mapping, parsers, readers."""
import csv
import io
import re

# Column name mapping: French variants -> internal field names
COLUMN_MAP: dict[str, str] = {
    # last_name
    "nom": "last_name",
    "nom de famille": "last_name",
    "last_name": "last_name",
    "last name": "last_name",
    "lastname": "last_name",
    # first_name
    "prenom": "first_name",
    "prénom": "first_name",
    "first_name": "first_name",
    "first name": "first_name",
    "firstname": "first_name",
    # email
    "email": "email",
    "e mail": "email",
    "e-mail": "email",
    "mail": "email",
    "adresse email": "email",
    "adresse e mail": "email",
    "adresse e-mail": "email",
    # phone
    "telephone": "phone",
    "téléphone": "phone",
    "tel": "phone",
    "tél": "phone",
    "phone": "phone",
    "portable": "phone",
    "mobile": "phone",
    # birth_date
    "date de naissance": "birth_date",
    "date naissance": "birth_date",
    "date_naissance": "birth_date",
    "naissance": "birth_date",
    "birth_date": "birth_date",
    "birth date": "birth_date",
    # address
    "adresse": "address",
    "address": "address",
    # city
    "ville": "city",
    "city": "city",
    # postal_code
    "code postal": "postal_code",
    "code_postal": "postal_code",
    "cp": "postal_code",
    "postal_code": "postal_code",
    "postal code": "postal_code",
    # social_security_number
    "n° sécu": "social_security_number",
    "n° secu": "social_security_number",
    "numero secu": "social_security_number",
    "numéro sécu": "social_security_number",
    "nir": "social_security_number",
    "securite sociale": "social_security_number",
    "sécurité sociale": "social_security_number",
    "social_security_number": "social_security_number",
    "social security number": "social_security_number",
    # notes
    "notes": "notes",
    "commentaire": "notes",
    "commentaires": "notes",
}

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
PHONE_RE = re.compile(r"^[\d\s\+\-\.\(\)]{6,20}$")


def detect_delimiter(sample: str) -> str:
    """Auto-detect CSV delimiter from a text sample."""
    for delim in [";", ",", "\t"]:
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample, delimiters=delim)
            return dialect.delimiter
        except csv.Error:
            continue
    counts = {d: sample.count(d) for d in [";", ",", "\t"]}
    return max(counts, key=counts.get)  # type: ignore[arg-type]


def normalize_column(name: str) -> str:
    """Normalize a column header for mapping lookup."""
    return name.strip().lower().replace("_", " ").replace("-", " ").strip()


def parse_date(val: str) -> str | None:
    """Try to parse a date string into YYYY-MM-DD format."""
    from datetime import datetime as _dt

    val = val.strip()
    if not val:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%y"):
        try:
            return _dt.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def read_excel_rows(file_content: bytes) -> list[dict[str, str]]:
    """Read rows from an Excel (.xlsx) file and return list of dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [str(h).strip() if h else "" for h in header_row]
    result: list[dict[str, str]] = []
    for row in rows_iter:
        row_dict: dict[str, str] = {}
        for idx, cell_val in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = str(cell_val).strip() if cell_val is not None else ""
        result.append(row_dict)
    wb.close()
    return result
