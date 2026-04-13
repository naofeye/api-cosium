"""Styles openpyxl reutilisables pour les exports Excel."""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def apply_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Ecrit la ligne d'entete (row 1) avec style standard OptiFlow."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def set_column_widths(ws: Worksheet, widths: list[int]) -> None:
    """Largeur des colonnes A, B, C... selon la liste fournie."""
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
