"""Export Excel — balance clients (factures impayees aggregees par client)."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.core.logging import get_logger, log_operation
from app.services._export_styles import THIN_BORDER, apply_header_row, set_column_widths

logger = get_logger("export_xlsx_balance")


@log_operation("export_balance_clients_xlsx")
def export_balance_clients_xlsx(
    db: Session,
    tenant_id: int,
    date_from=None,
    date_to=None,
) -> bytes:
    """Generate an Excel balance report of outstanding client debts."""
    from app.services.export_pdf import _get_balance_rows

    rows = _get_balance_rows(db, tenant_id, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Clients"

    headers = ["Client", "Nb Factures", "Total Facture (EUR)", "Total Impaye (EUR)", "Derniere Facture"]
    apply_header_row(ws, headers)

    total_facture = 0.0
    total_impaye = 0.0
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["client"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=row["nb_factures"]).border = THIN_BORDER
        c_fact = ws.cell(row=row_idx, column=3, value=row["total_facture"])
        c_fact.number_format = "#,##0.00"
        c_fact.border = THIN_BORDER
        c_imp = ws.cell(row=row_idx, column=4, value=row["total_impaye"])
        c_imp.number_format = "#,##0.00"
        c_imp.border = THIN_BORDER
        date_val = row["derniere_facture"].strftime("%d/%m/%Y") if row["derniere_facture"] else ""
        ws.cell(row=row_idx, column=5, value=date_val).border = THIN_BORDER
        total_facture += row["total_facture"]
        total_impaye += row["total_impaye"]

    # Total row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(rows)).font = Font(bold=True)
    c_tf = ws.cell(row=total_row, column=3, value=round(total_facture, 2))
    c_tf.number_format = "#,##0.00"
    c_tf.font = Font(bold=True)
    c_ti = ws.cell(row=total_row, column=4, value=round(total_impaye, 2))
    c_ti.number_format = "#,##0.00"
    c_ti.font = Font(bold=True)

    set_column_widths(ws, [35, 14, 20, 20, 18])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("balance_clients_xlsx_exported", tenant_id=tenant_id, rows=len(rows))
    return output.getvalue()
